import openpyxl

book = openpyxl.load_workbook(r"D:\Data\testData.xlsx")
sheet = book.active
cell = sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=2,column=2).value = "JyotirmY"
cell = sheet.cell(row=2,column=2)
Dict ={}
# print(sheet['A5'].value)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "TestCase2":
        for j in range(1,sheet.max_column+1):
            Dict[sheet.cell(row= 1,column=j).value]=sheet.cell(row= i,column=j).value


print(Dict)